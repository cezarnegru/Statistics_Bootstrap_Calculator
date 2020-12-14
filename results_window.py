#!/usr/bin/env python3

import openpyxl
import random
import tkinter as tk
import main_window
import os

global mean_of_means
global new_mean_list
global mean_per_list

def main():
    default_entry_data_text = ['File path:','Spreadsheet:','Column:','Resamples:']
    data_entries = []
    entry_data_txt = 'entry_data.txt'

    if os.path.isfile(entry_data_txt):
        f = open(entry_data_txt, 'r')
        for line in f.readlines():
            data_entries.append(line.strip())
        f.close()

    for i in range(len(data_entries)):
        data_entries[i] = data_entries[i].split(default_entry_data_text[i])

    data = []
    for i in range(len(data_entries)):
        for j in range(len(data_entries[i])):
            if j == 1:
                data.append(data_entries[i][j])

    # Path/File
    wk = openpyxl.load_workbook(str(data[0]))
    # Sheet
    sh = wk[str(data[1])]
    # Column
    column = data[2]
    # Resamples
    num_of_resamples = int(data[3])
    
    original_list = []
    if column.isdigit():
        for i in range(1,sh.max_row+1):
            if sh.cell(i,int(column)).value == None:
                continue
            elif str(sh.cell(i,int(column)).value).isdigit():
                original_list.append(int(sh.cell(i,int(column)).value))
            else:
                continue
    elif column.isalpha():
        for r in sh[f'{column}{1}':f'{column}{sh.max_row+1}']:
            for c in r:
                if c.value == None:
                    continue
                elif str(c.value).isdigit():
                    original_list.append(c.value)
                else:
                    continue

    # Create resample from the original list
    def new_resample(original_sample):
        newResample = []
        while len(newResample) != len(original_sample):
            i = random.randint(0,len(original_sample)-1)
            newResample.append(original_sample[i])
        return newResample

    # Calculate mean of a given list
    def resample_mean(resample):
        mean = 0
        for score in resample:
            mean += score
        return mean/len(resample)

    # Calculate bootstrap (mean of means) from all resampled lists
    def bootstrap(original_sample, num_of_resamples, decimal_length=2):
        means = []
        d_l = decimal_length
        for res_mean in range(0,num_of_resamples):
            new_res = new_resample(original_sample)
#            print(new_res)
            means.append(float('%.2f'%resample_mean(new_res)))
        return (f'%.{d_l}f' %(resample_mean(means)))
    
    mean = str(bootstrap(original_list,num_of_resamples,2))
    
    results_window = tk.Tk()
    results_window.title('Results')
    results_window.geometry('550x280')
    results_window.configure(bg=main_window.window_bg)
    results_window.resizable(width=True,height=False)

    # LABEL = SPACE AT THE BEGINNING '\t'*2
    label_space = tk.Label(results_window,text='\tRESULTS\n\t_________\n', fg=main_window.label_fg, bg=main_window.window_bg, font=main_window.button_font, pady=main_window.button_pady, bd=main_window.button_bd, highlightbackground=main_window.window_bg)
    label_space.grid(row=0,column=0,columnspan=3)
    # LABEL = FILE NAME
    label_file_read = tk.Label(results_window,text='\t             File: '+data[0],fg=main_window.label_fg, bg=main_window.window_bg, font=main_window.button_font, pady=main_window.button_pady, bd=main_window.button_bd, highlightbackground=main_window.window_bg)
    label_file_read.grid(row=1,column=1)
    # LABEL = SHEET NAME
    label_sheet = tk.Label(results_window,text='\t          Sheet: '+data[1],fg=main_window.label_fg, bg=main_window.window_bg, font=main_window.button_font, pady=main_window.button_pady, bd=main_window.button_bd, highlightbackground=main_window.window_bg)
    label_sheet.grid(row=2,column=1)
    # LABEL = NUMBER OF RESAMPLES
    label_num_of_resamples = tk.Label(results_window,text='Number of resamples: '+data[3],fg=main_window.label_fg, bg=main_window.window_bg, font=main_window.button_font, pady=main_window.button_pady, bd=main_window.button_bd, highlightbackground=main_window.window_bg)
    label_num_of_resamples.grid(row=3,column=1)
    # LABEL = MEAN
    label_mean = tk.Label(results_window,text='\t          Mean: '+mean,fg=main_window.label_fg, bg=main_window.window_bg, font=main_window.button_font, pady=main_window.button_pady, bd=main_window.button_bd, highlightbackground=main_window.window_bg)
    label_mean.grid(row=4,column=1)
    # BUTTON = EXIT 
    button_exit = tk.Button(results_window,text='Exit', fg=main_window.button_fg, bg=main_window.button_bg, font=main_window.button_font, bd=main_window.button_bd, padx=30,pady=main_window.button_pady, highlightbackground=main_window.window_bg,command=results_window.destroy)
    button_exit.grid(row=6,column=1)

    results_window.mainloop()

if __name__ == '__main__':
    main()
